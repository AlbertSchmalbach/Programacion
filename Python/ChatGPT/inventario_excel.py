import openpyxl

def agregar_producto(archivo):
    producto = input("Ingrese el nombre del producto: ")
    cantidad = int(input("Ingrese la cantidad: "))
    precio_unitario = float(input("Ingrese el precio unitario: "))
    
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active
    
    ultima_fila = hoja.max_row + 1
    hoja.cell(row=ultima_fila, column=1, value=producto)
    hoja.cell(row=ultima_fila, column=2, value=cantidad)
    hoja.cell(row=ultima_fila, column=3, value=precio_unitario)
    
    wb.save(archivo)
    wb.close()

def mostrar_inventario(archivo):
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active
    
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        producto, cantidad, precio_unitario = fila
        total = cantidad * precio_unitario
        print("Producto:", producto)
        print("Cantidad:", cantidad)
        print("Precio Unitario:", precio_unitario)
        print("Total:", total)
        print("----------------------")
    
    wb.close()

archivo_excel = r"C:\Users\KHAdmin\Documents\CarpetaProtegida\Albert\Archivos\Script_Python\chatGPT\inventario.xlsx"

while True:
    print("1. Agregar Producto")
    print("2. Mostrar Inventario")
    print("3. Salir")
    opcion = input("Seleccione una opción: ")
    
    if opcion == "1":
        agregar_producto(archivo_excel)
    elif opcion == "2":
        mostrar_inventario(archivo_excel)
    elif opcion == "3":
        break
    else:
        print("Opción no válida")

print("¡Hasta luego!")
