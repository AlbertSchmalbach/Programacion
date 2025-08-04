import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import sys, os

application_path = os.path.dirname(sys.executable)

def automatizar_excel(nombre_archivo):
    
    """" Input sales_mes.xlsx | Output report_mes.xlsx """
    
    # leer archivo excel
    archivo_excel = pd.read_excel(f'{application_path}/{nombre_archivo}')
    # print(archivo_excel[['Gender', 'Product line', 'Total']])

    # Crear tabla pivote
    tabla_pivote = pd.pivot_table(data=archivo_excel, index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    # print(tabla_pivote)
    
    mes_extension = nombre_archivo.split('_')[1]

    # General archivo excel con los resultados de la tabla pivote
    tabla_pivote.to_excel(f'ventas_{mes_extension}', startrow=4, sheet_name='Report')

    # Leer libro excel con la propiedad de openpyxl load_workbook
    wb = load_workbook(f'ventas_{mes_extension}')

    # Llamar a la pestaña Report
    hoja_report = wb['Report']

    # valores minimos y maximos de filas y columnas
    min_colunm = wb.active.min_column
    max_colunm =wb.active.max_column
    min_fila = wb.active.min_row
    max_fila = wb.active.max_row

    # grafico
    barchart = BarChart()

    data = Reference(hoja_report, min_col=min_colunm+1, max_col=max_colunm, min_row=min_fila, max_row=max_fila)
    categorias = Reference(hoja_report, min_col=min_colunm, max_col=min_colunm, min_row=min_fila+1, max_row=max_fila)

    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categorias)

    hoja_report.add_chart(barchart, 'B12')
    barchart.title = 'Ventas'
    barchart.style = 3

    for c in ['B', 'C', 'D', 'E', 'F', 'G']:
        hoja_report[f'{c}8'] = f'=SUM({c}6:{c}7)'
        hoja_report[f'{c}8'].style = 'Currency'



    hoja_report['A1'] = "REPORTE"
    mes = mes_extension.split('.')[0]
    hoja_report['A2'] = mes


    hoja_report['A1'].font = Font('Arial', bold=True, size=30)
    hoja_report['A2'].font = Font('Arial', bold=True, size=25)



    wb.save(f'ventas_{mes_extension}')
    
    return

automatizar_excel('supermarket_sales.xlsx')
