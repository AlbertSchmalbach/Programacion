import openpyxl

wb = openpyxl.load_workbook("planilla.xlsx")



hoja = wb.active

wb.create_sheet("Miscelaneo")

hoja4 = wb["Miscelaneo"]
hoja4.title = "Biografia"
# print(wb.sheetnames)

# print(hoja.max_row, hoja.max_column)

celda = hoja["A1"]
celda.value = "Nombre completo"
# print(celda.value)

celda2 = hoja.cell(row=2, column=1)
print(celda2.value,
      celda2.row,
      celda.column,
      celda2.coordinate
)